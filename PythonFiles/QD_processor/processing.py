import openpyxl as xl
import os


def create_workbook(wb_path):
    wb = xl.Workbook()
    wb.save(filename=wb_path)


def write_txt_to_wb(txt_path, wb_path):
    data_list = []
    with open(txt_path, 'r') as txt:
        for line in txt:
            data_list.append(line)
    print(data_list)
    wb = xl.load_workbook(filename=wb_path)
    for sheetname in wb.sheetnames:
        wb.remove(wb[sheetname])
    wb.create_sheet('I-V Data')
    sheet = wb.active
    for value in data_list:
        data = value.strip().split('\t')
        mod_data = []
        for datum in data:
            mod_data.append(datum.strip())
        sheet.append(mod_data)
    wb.save(filename=wb_path)


def calculate(wb_path, size):
    j_dark = {}
    j_light = {}
    eqe = {}
    if size == 0:
        area = 0.0004
    elif size == 1:
        area = 0.01
    elif size == 2:
        area = 0.04
    else:
        area = 0.09
    wb = xl.load_workbook(filename=wb_path)
    sheet = wb['I-V Data']
    j_dark['-0.5V'] = float(sheet['C127'].value) * 1e9 / area
    j_dark['-2V'] = float(sheet['C202'].value) * 1e9 / area
    j_light['-0.5V'] = float(sheet['C328'].value) * 1e9 / area
    j_light['-2V'] = float(sheet['C403'].value) * 1e9 / area
    eqe['-0.5V'] = (j_light['-0.5V'] - j_dark['-0.5V']) * 1240 / 1e9 / 0.001 / 970
    eqe['-2V'] = (j_light['-2V'] - j_dark['-2V']) * 1240 / 1e9 / 0.001 / 970
    return {'Jdark': j_dark, 'EQE': eqe}


def write_result_to_sheet(wb_path, result):
    wb = xl.load_workbook(filename=wb_path)
    sheet = wb['I-V Data']
    sheet['E1'].value = '-0.5V'
    sheet['H1'].value = '-2V'
    sheet['E2'].value = 'Jdark (nA/cm2)'
    sheet['F2'].value = 'EQE'
    sheet['H2'].value = 'Jdark (nA/cm2)'
    sheet['I2'].value = 'EQE'
    sheet['E3'].value = result['Jdark']['-0.5V']
    sheet['F3'].value = result['EQE']['-0.5V']
    sheet['H3'].value = result['Jdark']['-2V']
    sheet['I3'].value = result['EQE']['-2V']
    wb.save(filename=wb_path)


def process(txt_path, size):
    wb_path = txt_path.replace('.txt', '.xlsx')
    try:
        if not os.path.exists(wb_path):
            create_workbook(wb_path)
        write_txt_to_wb(txt_path, wb_path)
        result = calculate(wb_path, size)
        write_result_to_sheet(wb_path, result)
    except PermissionError:
        result = {'Jdark': {'-0.5V': 0, '-2V': 0}, 'EQE': {'-0.5V': 0, '-2V': 0}}
    return result
