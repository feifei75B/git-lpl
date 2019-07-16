import openpyxl as xl
import os


class ExcelFile(object):

    def __init__(self, path, sheetname):
        self.path = path
        self.wb = self.get_workbook()
        self.sheet = self.get_sheet(sheetname)
        self.save()

    def get_workbook(self):
        if os.path.exists(self.path):
            return xl.load_workbook(self.path)
        else:
            return xl.Workbook()

    def get_sheet(self, sheetname):
        if sheetname in self.wb.sheetnames:
            return self.wb[sheetname]
        else:
            self.wb.create_sheet(sheetname)
        return self.wb.active

    def write_data(self, data, position):
        self.sheet[position].value = data

    def write_data_list(self, data_list):
        self.sheet.append(data_list)

    def save(self):
        self.wb.save(self.path)


class PointIVXlsx(ExcelFile):

    def __init__(self, path):
        super().__init__(path, 'Data')
        self.max_column = 0
        self.save()

    def get_sheet(self, sheetname):
        for sheetname in self.wb.sheetnames:
            self.wb.remove(self.wb[sheetname])
        self.wb.create_sheet(sheetname)
        return self.wb.active

    def set_max_column(self, column):
        self.max_column = column

    def write_data_dict(self, data, row_o, column_o):
        row = row_o
        if column_o > self.max_column:
            self.max_column = column_o
        if isinstance(data, dict):
            for key, value in data.items():
                row = row_o
                self.sheet.cell(row=row, column=self.max_column).value = key
                row += 1
                self.write_data_dict(value, row, self.max_column)
        else:
            self.sheet.cell(row=row, column=self.max_column).value = data
            self.max_column += 1


class PointIVXlsxs(ExcelFile):

    def __init__(self, dir_path, device_id):
        super().__init__(dir_path + '\\' + device_id + '.xlsx', device_id)
        self.save()

    def get_sheet(self, sheetname):
        for sheetname in self.wb.sheetnames:
            self.wb.remove(self.wb[sheetname])
        self.wb.create_sheet(sheetname)
        return self.wb.active

    def write_show_data(self, show_data):
        for data_list in show_data:
            self.sheet.append(data_list)
